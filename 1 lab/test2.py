import openpyxl
from openpyxl.utils import get_column_letter
import os


def prepare_excel_for_solver():
    # Создаем новую книгу и выбираем активный лист
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Solver_Data"

    print("Заполнение Excel данными...")

    # --- Заполнение данных ---
    data = {
        "A1": "Тип корма", "B1": "Regular (x_R)", "C1": "Extra (x_E)", "D1": "Puppy Delite (x_P)", "F1": "Расход", "G1": "Знак", "H1": "Запас",
        "A2": "Прибыль", "B2": 20, "C2": 18, "D2": 25,
        "A3": "План (кг)", "B3": 0, "C3": 0, "D3": 0,  # Начальные значения
        "A4": "Общая прибыль", "E4": "=SUMPRODUCT(B2:D2, B3:D3)",
        "A6": "Ингредиент A", "B6": 1/3, "C6": 0.5, "D6": 0, "F6": "=SUMPRODUCT(B6:D6, $B$3:$D$3)", "G6": "<=", "H6": 1900,
        "A7": "Ингредиент B", "B7": 1/3, "C7": 0.25, "D7": 0.1, "F7": "=SUMPRODUCT(B7:D7, $B$3:$D$3)", "G7": "<=", "H7": 1100,
        "A8": "Ингредиент C", "B8": 1/3, "C8": 0.25, "D8": 0.9, "F8": "=SUMPRODUCT(B8:D8, $B$3:$D$3)", "G8": "<=", "H8": 1000,
    }

    for cell, value in data.items():
        worksheet[cell] = value

    # Сделаем столбцы пошире для наглядности
    for i in range(1, 9):
        worksheet.column_dimensions[get_column_letter(i)].width = 18

    # --- Сохранение файла ---
    filename = "solver_task.xlsx"
    workbook.save(filename)

    # Получаем полный путь к файлу
    full_path = os.path.abspath(filename)

    print("\n--- ГОТОВО! ---")
    print(f"Файл '{filename}' создан и заполнен.")
    print(f"Он находится здесь: {full_path}")
    print("\n--- Ваши дальнейшие действия: ---")
    print("1. Откройте этот файл в Excel.")
    print("2. Перейдите на вкладку 'Данные' и нажмите 'Поиск решения'.")
    print("3. Настройте параметры:")
    print("   - Оптимизировать целевую функцию: $E$4")
    print("   - До: Максимум")
    print("   - Изменяя ячейки переменных: $B$3:$D$3")
    print("   - В соответствии с ограничениями:")
    print("     $F$6 <= $H$6")
    print("     $F$7 <= $H$7")
    print("     $F$8 <= $H$8")
    print("   - Поставьте галочку 'Сделать переменные без ограничений неотрицательными'.")
    print("   - Метод решения: 'Симплекс-метод ЛП'.")
    print("4. Нажмите 'Найти решение'.")
    print("5. В появившемся окне выберите 'Устойчивость' в списке отчетов и нажмите ОК.")


# Запускаем функцию
prepare_excel_for_solver()
